import csv
import io
import logging
from datetime import datetime
from typing import Optional, List

from sqlalchemy.orm import Session

from app.models.monitor import MonitorAccount, Project
from app.schemas.account import (
    AccountCreate,
    BatchAccountResult,
    BatchAccountResultItem,
    BatchActionRequest,
)
from app.services.monitor_service import create_account, check_account, get_account

logger = logging.getLogger(__name__)

# CSV 导出字段 - 包含所有重要账号信息和24小时对比
_EXPORT_FIELDS = [
    "username",           # 用户名
    "nickname",           # 昵称
    "tiktok_id",         # TikTok ID
    "sec_uid",           # Sec UID
    "account_created_at", # 账号注册时间
    "region",            # 地区/国家
    "follower_count",    # 粉丝数
    "follower_count_24h", # 24小时前粉丝数
    "follower_change_24h", # 24小时粉丝变化
    "following_count",   # 关注数
    "following_count_24h", # 24小时前关注数
    "following_change_24h", # 24小时关注变化
    "like_count",        # 点赞数
    "like_count_24h",    # 24小时前点赞数
    "like_change_24h",   # 24小时点赞变化
    "video_count",       # 视频数
    "video_count_24h",   # 24小时前视频数
    "video_change_24h",  # 24小时视频变化
    "bio",               # 简介
    "avatar_url",        # 头像URL
    "is_active",         # 是否启用
    "monitor_interval",  # 监控间隔（秒）
    "use_proxy",         # 是否使用代理
    "enable_video_monitoring",  # 是否启用视频监控
    "last_checked_at",   # 最后检查时间
    "created_at",        # 创建时间
    "updated_at",        # 更新时间
]


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

def _get_export_accounts(db: Session, project_id: Optional[int]) -> List[MonitorAccount]:
    query = db.query(MonitorAccount)
    if project_id is not None:
        query = query.filter(MonitorAccount.project_id == project_id)
    return query.order_by(MonitorAccount.id).all()


def _get_24h_ago_data(db: Session, account_id: int) -> Optional[dict]:
    """获取24小时前的历史数据"""
    from app.models.monitor import MonitorHistory
    from datetime import timedelta
    
    # 计算24小时前的时间点
    time_24h_ago = datetime.utcnow() - timedelta(hours=24)
    
    # 查找最接近24小时前的历史记录
    history = (
        db.query(MonitorHistory)
        .filter(
            MonitorHistory.account_id == account_id,
            MonitorHistory.checked_at <= time_24h_ago,
            MonitorHistory.check_status == "success"
        )
        .order_by(MonitorHistory.checked_at.desc())
        .first()
    )
    
    if history:
        return {
            "follower_count": history.follower_count,
            "following_count": history.following_count,
            "like_count": history.like_count,
            "video_count": history.video_count,
        }
    return None


def _format_value(account: MonitorAccount, field: str) -> str:
    value = getattr(account, field, "")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if value is None:
        return ""
    return str(value)


def export_accounts_csv(db: Session, project_id: Optional[int] = None) -> str:
    """导出账号列表为 CSV 字符串，包含24小时对比数据。"""
    accounts = _get_export_accounts(db, project_id)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=_EXPORT_FIELDS)
    writer.writeheader()
    for acc in accounts:
        data_24h_ago = _get_24h_ago_data(db, acc.id)
        row = {}
        for f in _EXPORT_FIELDS:
            if f.endswith("_24h") and not f.endswith("_change_24h"):
                field_name = f.replace("_24h", "")
                row[f] = data_24h_ago.get(field_name, "") if data_24h_ago else "-"
            elif f.endswith("_change_24h"):
                field_name = f.replace("_change_24h", "")
                current = getattr(acc, field_name, 0) or 0
                if data_24h_ago:
                    old = data_24h_ago.get(field_name, 0) or 0
                    change = current - old
                    row[f] = f"+{change}" if change > 0 else str(change)
                else:
                    row[f] = "-"
            else:
                row[f] = _format_value(acc, f)
        writer.writerow(row)
    return output.getvalue()


def export_accounts_excel(db: Session, project_id: Optional[int] = None) -> bytes:
    """导出账号列表为 Excel (xlsx) bytes，包含24小时数据对比。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export. Install it via 'pip install openpyxl'.")

    accounts = _get_export_accounts(db, project_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts"
    
    # 中文表头映射
    header_map = {
        "username": "用户名",
        "nickname": "昵称",
        "tiktok_id": "TikTok ID",
        "sec_uid": "Sec UID",
        "account_created_at": "账号注册时间",
        "region": "地区/国家",
        "follower_count": "粉丝数",
        "follower_count_24h": "24h前粉丝数",
        "follower_change_24h": "24h粉丝变化",
        "following_count": "关注数",
        "following_count_24h": "24h前关注数",
        "following_change_24h": "24h关注变化",
        "like_count": "点赞数",
        "like_count_24h": "24h前点赞数",
        "like_change_24h": "24h点赞变化",
        "video_count": "视频数",
        "video_count_24h": "24h前视频数",
        "video_change_24h": "24h视频变化",
        "bio": "简介",
        "avatar_url": "头像URL",
        "is_active": "启用状态",
        "monitor_interval": "监控间隔(分钟)",
        "use_proxy": "使用代理",
        "enable_video_monitoring": "视频监控",
        "last_checked_at": "最后检查时间",
        "created_at": "创建时间",
        "updated_at": "更新时间",
    }
    
    # 写入中文表头
    header_row = [header_map.get(f, f) for f in _EXPORT_FIELDS]
    ws.append(header_row)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入数据行
    for acc in accounts:
        # 获取24小时前的数据
        data_24h_ago = _get_24h_ago_data(db, acc.id)
        
        row_data = []
        for f in _EXPORT_FIELDS:
            # 处理24小时对比字段
            if f.endswith("_24h") and not f.endswith("change_24h"):
                # 24小时前的数据
                field_name = f.replace("_24h", "")
                if data_24h_ago:
                    row_data.append(data_24h_ago.get(field_name, ""))
                else:
                    row_data.append("-")
            elif f.endswith("_change_24h"):
                # 24小时变化
                field_name = f.replace("_change_24h", "")
                current_value = getattr(acc, field_name, 0) or 0
                if data_24h_ago:
                    old_value = data_24h_ago.get(field_name, 0) or 0
                    change = current_value - old_value
                    if change > 0:
                        row_data.append(f"+{change}")
                    elif change < 0:
                        row_data.append(str(change))
                    else:
                        row_data.append("0")
                else:
                    row_data.append("-")
            else:
                # 常规字段
                value = getattr(acc, f, "")
                # 处理不同类型的值
                if isinstance(value, datetime):
                    row_data.append(value.strftime("%Y-%m-%d %H:%M:%S"))
                elif value is None or value == "":
                    row_data.append("")
                elif isinstance(value, bool):
                    row_data.append("是" if value else "否")
                elif f == "monitor_interval":
                    # 将秒转换为分钟显示
                    row_data.append(value // 60 if value else 0)
                elif isinstance(value, (int, float)):
                    row_data.append(value)
                else:
                    # 字符串类型，截断过长的内容
                    str_value = str(value)
                    if f in ["sec_uid", "avatar_url"] and len(str_value) > 50:
                        row_data.append(str_value[:50] + "...")
                    elif f == "bio" and len(str_value) > 100:
                        row_data.append(str_value[:100] + "...")
                    else:
                        row_data.append(str_value)
        ws.append(row_data)
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最大宽度50
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存到 BytesIO
    output = io.BytesIO()
    try:
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Import helpers
# ---------------------------------------------------------------------------

def _parse_usernames_from_text(text: str) -> List[str]:
    """从多行文本中提取用户名，去除空行和首尾空格。"""
    usernames = []
    for line in text.splitlines():
        username = line.strip().lstrip("@")
        if username:
            usernames.append(username)
    return usernames


async def import_accounts_from_text(
    db: Session,
    project_id: int,
    text: str,
    monitor_interval: int = 3600,
) -> BatchAccountResult:
    """解析多行文本批量创建账号，返回 BatchAccountResult。"""
    usernames = _parse_usernames_from_text(text)
    results: List[BatchAccountResultItem] = []
    success_count = 0
    duplicate_count = 0
    failed_count = 0

    for username in usernames:
        try:
            data = AccountCreate(
                project_id=project_id,
                username=username,
                monitor_interval=monitor_interval,
            )
            account = create_account(db, data)
            results.append(BatchAccountResultItem(username=username, status="success"))
            success_count += 1
        except ValueError as e:
            err = str(e)
            if "already exists" in err:
                results.append(
                    BatchAccountResultItem(username=username, status="duplicate", reason=err)
                )
                duplicate_count += 1
            else:
                results.append(
                    BatchAccountResultItem(username=username, status="failed", reason=err)
                )
                failed_count += 1
        except Exception as e:
            results.append(
                BatchAccountResultItem(username=username, status="failed", reason=str(e)[:200])
            )
            failed_count += 1

    return BatchAccountResult(
        total=len(usernames),
        success=success_count,
        duplicates=duplicate_count,
        failed=failed_count,
        results=results,
    )


async def import_accounts_from_file(
    db: Session,
    project_id: int,
    file_content: bytes,
    filename: str,
    monitor_interval: int = 3600,
) -> BatchAccountResult:
    """解析 CSV 或 Excel 文件，提取 username 列，调用 import_accounts_from_text。"""
    lower_name = filename.lower()

    if lower_name.endswith(".csv"):
        usernames = _extract_usernames_from_csv(file_content)
    elif lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
        usernames = _extract_usernames_from_excel(file_content)
    else:
        raise ValueError(f"Unsupported file format: {filename}. Only CSV and Excel files are supported.")

    text = "\n".join(usernames)
    return await import_accounts_from_text(db, project_id, text, monitor_interval)


def _extract_usernames_from_csv(content: bytes) -> List[str]:
    """从 CSV bytes 中提取 username 列，自动跳过表头。"""
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))

    # 检查是否有 username 列
    if reader.fieldnames is None:
        raise ValueError("CSV file appears to be empty")

    fieldnames_lower = [f.strip().lower() for f in reader.fieldnames]
    if "username" not in fieldnames_lower:
        raise ValueError("CSV file must contain a 'username' column")

    # 找到实际列名（大小写不敏感）
    username_col = reader.fieldnames[
        fieldnames_lower.index("username")
    ]

    # Common header values to skip (case-insensitive)
    header_values = {"username", "user", "name", "account"}

    usernames = []
    for row in reader:
        val = (row.get(username_col) or "").strip().lstrip("@")
        # Skip empty values and common header values
        if val and val.lower() not in header_values:
            usernames.append(val)
    return usernames


def _extract_usernames_from_excel(content: bytes) -> List[str]:
    """从 Excel bytes 中提取 username 列，自动跳过表头。"""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel import.")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file appears to be empty")

    # 第一行作为表头
    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    if "username" not in header:
        raise ValueError("Excel file must contain a 'username' column")

    col_idx = header.index("username")
    usernames = []
    for row in rows[1:]:
        if col_idx < len(row):
            val = str(row[col_idx]).strip().lstrip("@") if row[col_idx] is not None else ""
            if val and val.lower() != "none":
                usernames.append(val)
    return usernames


# ---------------------------------------------------------------------------
# Batch actions
# ---------------------------------------------------------------------------

async def batch_action(db: Session, data: BatchActionRequest) -> dict:
    """批量操作：enable / disable / delete / move。"""
    action = data.action
    account_ids = data.account_ids

    if not account_ids:
        return {"affected": 0}

    accounts = (
        db.query(MonitorAccount)
        .filter(MonitorAccount.id.in_(account_ids))
        .all()
    )

    if not accounts:
        return {"affected": 0}

    if action == "enable":
        for acc in accounts:
            acc.is_active = True
        db.commit()

    elif action == "disable":
        for acc in accounts:
            acc.is_active = False
        db.commit()

    elif action == "delete":
        for acc in accounts:
            db.delete(acc)
        db.commit()

    elif action == "move":
        target_project_id = data.target_project_id
        if target_project_id is None:
            raise ValueError("target_project_id is required for 'move' action")
        # 验证目标项目存在
        target_project = db.query(Project).filter(Project.id == target_project_id).first()
        if not target_project:
            raise LookupError(f"Target project {target_project_id} not found")
        for acc in accounts:
            acc.project_id = target_project_id
        db.commit()

    else:
        raise ValueError(f"Unknown action: '{action}'. Valid actions: enable, disable, delete, move")

    return {"affected": len(accounts)}
